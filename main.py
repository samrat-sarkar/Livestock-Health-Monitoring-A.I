from chatterbot import ChatBot
from chatterbot.trainers import ListTrainer
from openpyxl import load_workbook
import random

chatbot = ChatBot("Rubicon AI")

wb = load_workbook(filename='conversation.xlsx')
ws = wb.active
conversation = []
for row in ws.iter_rows(min_row=1, values_only=True):
    conversation.append(row[0])

trainer = ListTrainer(chatbot)
trainer.train(conversation)

Low_Confidence_msg = [
    "I'm sorry, I'm not sure what you mean. Could you please provide more details?",
    "I'm sorry, I couldn't understand your question. Can you try rephrasing it?",
    "I'm afraid I didn't catch that. Could you please repeat your question?",
    "I'm sorry, I didn't get that. Can you please ask a different question?",
    "I'm sorry, I'm not sure how to respond to that. Can you please try a different question?",
    "I'm sorry, I'm not sure what you're asking. Can you please clarify?",
    "I'm sorry, I don't have an answer to that. Could you please ask a different question?",
    "I'm sorry, I didn't understand your question. Could you please provide more context?",
    "I'm sorry, I'm not sure how to respond to that. Can you please ask a simpler question?",
    "I'm sorry, I didn't understand your question. Can you please try rephrasing it in a different way?"
]
print("Rubicon AI (Confidence: 1.0 ) >> How may I assist you ?")
while True:
    question = input("You >> ")
    response = chatbot.get_response(question)
    if response.confidence == 0:
        i = random.randint(0, 9)
        print(Low_Confidence_msg[i])
    else:
        print("Rubicon AI (Confidence:",response.confidence,") >>",response)

